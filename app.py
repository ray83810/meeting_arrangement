import os
import re
import datetime
import json
from flask import Flask, jsonify, request, render_template, send_from_directory
from werkzeug.utils import secure_filename


app = Flask(__name__, template_folder='templates', static_folder='static')

workspace_dir = os.path.dirname(os.path.abspath(__file__))

AGENT_MAP = {
    "alex chen": "Alex Chen",
    "amber wang": "Amber Wang",
    "evan liu": "Evan Liu",
    "howard chen": "Howard Chen",
    "jacky lee": "Jacky Lee",
    "jian kai ding": "Jian Kai Ding",
    "jiankai.ding": "Jian Kai Ding",
    "kai din": "Jian Kai Ding",
    "kai ding": "Jian Kai Ding",
    "molly song": "Molly Song",
    "rex liao": "Rex Liao",
    "rex laio": "Rex Liao",
    "sherry lin": "Sherry Lin"
}

SHIFTS_ALL = ["Alex Chen", "Amber Wang", "Evan Liu", "Howard Chen", "Jacky Lee", "Jian Kai Ding", "Molly Song", "Rex Liao", "Sherry Lin"]

# Cache for parsed monthly schedules
schedule_cache = {}

def normalize_name(name):
    if not name:
        return None
    name_clean = str(name).strip().lower().replace("  ", " ")
    return AGENT_MAP.get(name_clean, name_clean)

def parse_meal_hours(meal_val, start_hour):
    meal_hours = []
    if meal_val is None:
        return [(start_hour + 4) % 24]
    
    val_str = str(meal_val)
    matches = re.findall(r'(\d{1,2}):(\d{2})', val_str)
    if matches:
        for h_str, m_str in matches:
            meal_hours.append(int(h_str))
    else:
        if isinstance(meal_val, datetime.time):
            meal_hours.append(meal_val.hour)
        elif isinstance(meal_val, datetime.datetime):
            meal_hours.append(meal_val.hour)
        else:
            digits = re.findall(r'\d+', val_str)
            if digits:
                for d in digits:
                    val_int = int(d)
                    if 0 <= val_int < 24:
                        meal_hours.append(val_int)
    
    if not meal_hours:
        meal_hours = [(start_hour + 4) % 24]
    return list(set(meal_hours))

def get_agent_shift_hours(shift_val):
    if shift_val is None:
        return None
    
    val_str = str(shift_val).strip()
    if val_str in ['OFF', 'PTO', 'LOA', 'SL', 'FL', 'PTO-Half', 'Half-PTO', 'PL', 'AL', '0']:
        return None
        
    match = re.match(r'^(\d{1,2}):(\d{2})', val_str)
    if match:
        return int(match.group(1))
        
    if isinstance(shift_val, datetime.time):
        return shift_val.hour
    if isinstance(shift_val, datetime.datetime):
        return shift_val.hour
        
    try:
        val_int = int(float(val_str))
        if 0 <= val_int < 24:
            return val_int
    except ValueError:
        pass
        
    match_any = re.search(r'(\d{1,2}):(\d{2})', val_str)
    if match_any:
        return int(match_any.group(1))
        
    return None

def load_all_excel_schedules():
    global schedule_cache
    import openpyxl
    
    files = sorted([f for f in os.listdir(workspace_dir) if f.endswith('.xlsx') and f.startswith('2026') and f != '2026排班數據.xlsx'])
    
    for filename in files:
        month_code = filename.replace(".xlsx", "")
        path = os.path.join(workspace_dir, filename)
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            sheet_name = [s for s in wb.sheetnames if s.startswith('20')][0]
            sheet = wb[sheet_name]
            
            dates = []
            for c in range(10, sheet.max_column + 1):
                val = sheet.cell(row=2, column=c).value
                if val is None:
                    dates.append(None)
                    continue
                if isinstance(val, datetime.datetime):
                    dates.append(val.strftime("%Y-%m-%d"))
                elif isinstance(val, str):
                    dates.append(val.strip())
                else:
                    dates.append(str(val))
                    
            valid_dates = []
            col_mapping = {}
            for idx, d_str in enumerate(dates):
                col = 10 + idx
                if d_str and len(d_str) == 10 and d_str[4] == '-' and d_str[7] == '-':
                    valid_dates.append(d_str)
                    col_mapping[d_str] = col
                    
            agents_data = {}
            for r in range(3, sheet.max_row + 1):
                cell_a = sheet.cell(row=r, column=1).value
                name = normalize_name(cell_a)
                if name in SHIFTS_ALL:
                    default_shift = sheet.cell(row=r, column=2).value
                    meal_val = sheet.cell(row=r, column=4).value
                    
                    daily_schedule = {}
                    for d_str in valid_dates:
                        col = col_mapping[d_str]
                        cell_val = sheet.cell(row=r, column=col).value
                        daily_schedule[d_str] = cell_val
                        
                    agents_data[name] = {
                        "default_shift": default_shift,
                        "meal": meal_val,
                        "schedule": daily_schedule
                    }
            wb.close()
            schedule_cache[month_code] = {
                "dates": valid_dates,
                "agents": agents_data
            }
            print(f"Successfully loaded {month_code}: {len(agents_data)} agents, {len(valid_dates)} dates.")
        except Exception as e:
            print(f"Error loading {filename}: {e}")

# Load schedules on start
load_all_excel_schedules()

@app.route('/')
def index():
    return send_from_directory(workspace_dir, 'index.html')

@app.route('/api/init', methods=['GET'])
def get_init_data():
    months = []
    for month_code in sorted(schedule_cache.keys()):
        year = month_code[:4]
        month = month_code[4:]
        months.append({
            "code": month_code,
            "label": f"{year}年{month}月"
        })
    return jsonify({
        "months": months,
        "agents": SHIFTS_ALL
    })

@app.route('/api/schedule', methods=['POST'])
def schedule():
    data = request.json or {}
    month_code = data.get("month")
    selected_agents = data.get("agents", [])
    course_count = int(data.get("course_count", 4))
    course_duration = int(data.get("course_duration", 2))
    min_coverage = int(data.get("min_coverage", 2))
    joint_class = bool(data.get("joint_class", False))
    
    if not month_code or month_code not in schedule_cache:
        return jsonify({"error": "找不到該月份的班表資料"}), 400
        
    if not selected_agents:
        return jsonify({"error": "請至少選擇一位客服人員"}), 400
        
    month_data = schedule_cache[month_code]
    dates = month_data["dates"]
    agents_data = month_data["agents"]
    
    online_matrix_orig = {d: {h: [] for h in range(24)} for d in dates}
    for agent_name in SHIFTS_ALL:
        agent_info = agents_data.get(agent_name)
        if not agent_info:
            continue
        schedule = agent_info["schedule"]
        meal_val = agent_info["meal"]
        for d_str, cell_val in schedule.items():
            start_hour = get_agent_shift_hours(cell_val)
            if start_hour is not None:
                shift_hours = [(start_hour + i) % 24 for i in range(9)]
                meal_hours = parse_meal_hours(meal_val, start_hour)
                working_hours = [h for h in shift_hours if h not in meal_hours]
                for h in working_hours:
                    online_matrix_orig[d_str][h].append(agent_name)
                    
    available_agents = {d: {h: set(online_matrix_orig[d][h]) for h in range(24)} for d in dates}
    
    scheduled_courses = {name: [] for name in selected_agents}
    failed_schedules = []
    
    if joint_class:
        # Multi-person joint meeting logic (all selected agents attend at the same time)
        for c_num in range(1, course_count + 1):
            best_slot = None
            best_score = -999
            best_sum_online = -999
            
            for d_str in dates:
                # Any selected agent already has a course scheduled on this day?
                day_has_course = False
                for name in selected_agents:
                    if any(c["date"] == d_str for c in scheduled_courses[name]):
                        day_has_course = True
                        break
                if day_has_course:
                    continue
                
                # Check working hours and contiguous slots for all selected agents
                working_hours_by_agent = {}
                valid_agents_on_day = True
                for name in selected_agents:
                    agent_info = agents_data[name]
                    schedule = agent_info["schedule"]
                    meal_val = agent_info["meal"]
                    cell_val = schedule.get(d_str)
                    start_hour = get_agent_shift_hours(cell_val)
                    if start_hour is None:
                        valid_agents_on_day = False
                        break
                    
                    shift_hours = [(start_hour + i) % 24 for i in range(9)]
                    meal_hours = parse_meal_hours(meal_val, start_hour)
                    working_hours = [h for h in shift_hours if h not in meal_hours]
                    working_hours_by_agent[name] = set(working_hours)
                
                if not valid_agents_on_day:
                    continue
                
                # Find hours where all agents are working on this day d_str
                common_working_hours = sorted(list(set.intersection(*working_hours_by_agent.values())))
                
                # Search contiguous slots in common_working_hours
                for idx in range(len(common_working_hours) - course_duration + 1):
                    candidate_hours = common_working_hours[idx : idx + course_duration]
                    is_contiguous = True
                    for k in range(len(candidate_hours) - 1):
                        if (candidate_hours[k+1] - candidate_hours[k]) % 24 != 1:
                            is_contiguous = False
                            break
                    if not is_contiguous:
                        continue
                    
                    min_rem_online = 999
                    sum_rem_online = 0
                    valid_slot = True
                    
                    # Verify if all agents are available (not already taking a course at these hours)
                    for h in candidate_hours:
                        for name in selected_agents:
                            if name not in available_agents[d_str][h]:
                                valid_slot = False
                                break
                        if not valid_slot:
                            break
                        
                        rem_online = len(available_agents[d_str][h] - set(selected_agents))
                        min_rem_online = min(min_rem_online, rem_online)
                        sum_rem_online += rem_online
                        
                    if not valid_slot:
                        continue
                    
                    # Score is min_rem_online
                    score = min_rem_online
                    
                    if best_slot is None or score > best_score or (abs(score - best_score) < 0.01 and sum_rem_online > best_sum_online):
                        best_slot = (d_str, candidate_hours[0], candidate_hours)
                        best_score = score
                        best_sum_online = sum_rem_online
            
            if best_slot:
                d_str, h_start, candidate_hours = best_slot
                min_remaining = int(best_score + 0.5)
                
                for name in selected_agents:
                    scheduled_courses[name].append({
                        "date": d_str,
                        "start_hour": h_start,
                        "end_hour": (h_start + course_duration) % 24 if (h_start + course_duration) % 24 != 0 else 24,
                        "duration": course_duration,
                        "coverage_at_hours": [len(available_agents[d_str][h] - set(selected_agents)) for h in candidate_hours],
                        "total_online_at_hours": [len(available_agents[d_str][h]) for h in candidate_hours],
                        "violated": min_remaining < min_coverage
                    })
                    for h in candidate_hours:
                        available_agents[d_str][h].remove(name)
            else:
                for name in selected_agents:
                    failed_schedules.append({
                        "agent": name,
                        "course_number": c_num
                    })
    else:
        # Original individual scheduling logic
        placement_requests = []
        for c_idx in range(course_count):
            for name in selected_agents:
                placement_requests.append((name, c_idx + 1))
                
        for name, c_num in placement_requests:
            agent_info = agents_data[name]
            schedule = agent_info["schedule"]
            meal_val = agent_info["meal"]
            
            best_slot = None
            best_score = -999
            best_sum_online = -999
            
            for d_str in dates:
                day_has_course = any(c["date"] == d_str for c in scheduled_courses[name])
                if day_has_course:
                    continue
                    
                cell_val = schedule.get(d_str)
                start_hour = get_agent_shift_hours(cell_val)
                if start_hour is None:
                    continue
                    
                shift_hours = [(start_hour + i) % 24 for i in range(9)]
                meal_hours = parse_meal_hours(meal_val, start_hour)
                working_hours = [h for h in shift_hours if h not in meal_hours]
                
                for idx in range(len(working_hours) - course_duration + 1):
                    candidate_hours = working_hours[idx : idx + course_duration]
                    is_contiguous = True
                    for k in range(len(candidate_hours) - 1):
                        if (candidate_hours[k+1] - candidate_hours[k]) % 24 != 1:
                            is_contiguous = False
                            break
                    if not is_contiguous:
                        continue
                        
                    min_other_online = 999
                    sum_other_online = 0
                    valid_slot = True
                    
                    for h in candidate_hours:
                        if name not in available_agents[d_str][h]:
                            valid_slot = False
                            break
                        other_online = len(available_agents[d_str][h] - {name})
                        min_other_online = min(min_other_online, other_online)
                        sum_other_online += other_online
                        
                    if not valid_slot:
                        continue
                        
                    courses_on_day = sum(1 for a in selected_agents for c in scheduled_courses[a] if c["date"] == d_str)
                    score = min_other_online - (courses_on_day * 0.05)
                    
                    if best_slot is None or score > best_score or (abs(score - best_score) < 0.01 and sum_other_online > best_sum_online):
                        best_slot = (d_str, candidate_hours[0], candidate_hours)
                        best_score = score
                        best_sum_online = sum_other_online
                        
            if best_slot:
                d_str, h_start, candidate_hours = best_slot
                min_remaining = int(best_score + 0.5)
                
                scheduled_courses[name].append({
                    "date": d_str,
                    "start_hour": h_start,
                    "end_hour": (h_start + course_duration) % 24 if (h_start + course_duration) % 24 != 0 else 24,
                    "duration": course_duration,
                    "coverage_at_hours": [len(available_agents[d_str][h] - {name}) for h in candidate_hours],
                    "total_online_at_hours": [len(available_agents[d_str][h]) for h in candidate_hours],
                    "violated": min_remaining < min_coverage
                })
                
                for h in candidate_hours:
                    available_agents[d_str][h].remove(name)
            else:
                failed_schedules.append({
                    "agent": name,
                    "course_number": c_num
                })
            
    coverage_timeline = {}
    for d_str in dates:
        day_coverage = []
        for h in range(24):
            before_count = len(online_matrix_orig[d_str][h])
            after_count = len(available_agents[d_str][h])
            day_coverage.append({
                "hour": h,
                "before": before_count,
                "after": after_count,
                "agents_before": list(online_matrix_orig[d_str][h]),
                "agents_after": list(available_agents[d_str][h])
            })
        coverage_timeline[d_str] = day_coverage
        
    return jsonify({
        "scheduled_courses": scheduled_courses,
        "failed_schedules": failed_schedules,
        "coverage_timeline": coverage_timeline,
        "dates": dates,
        "agents": SHIFTS_ALL
    })

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "沒有選擇檔案"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "沒有選擇檔案"}), 400
    
    if file and file.filename.endswith('.xlsx'):
        filename = secure_filename(file.filename)
        
        # Save to a temporary name to validate and parse
        import uuid
        temp_filename = f"temp_{uuid.uuid4().hex}.xlsx"
        temp_path = os.path.join(workspace_dir, temp_filename)
        file.save(temp_path)
        
        import openpyxl
        try:
            wb = openpyxl.load_workbook(temp_path, data_only=True)
            sheet_names = [s for s in wb.sheetnames if s.startswith('20')]
            if not sheet_names:
                wb.close()
                os.remove(temp_path)
                return jsonify({"error": "Excel 檔案中找不到以 '20' 開頭的班表分頁 (例如 202606)"}), 400
            
            sheet_name = sheet_names[0]
            # Try to get month code from filename first, e.g. 202608
            month_match = re.search(r'\d{6}', filename)
            if month_match:
                month_code = month_match.group(0)
            else:
                if len(sheet_name) == 6 and sheet_name.isdigit():
                    month_code = sheet_name
                else:
                    month_code = "".join(re.findall(r'\d+', sheet_name))
                    if not month_code or len(month_code) != 6:
                        month_code = filename.replace(".xlsx", "")
            
            # Verify dates row and columns
            sheet = wb[sheet_name]
            dates = []
            for c in range(10, sheet.max_column + 1):
                val = sheet.cell(row=2, column=c).value
                if isinstance(val, datetime.datetime):
                    dates.append(val.strftime("%Y-%m-%d"))
                elif isinstance(val, str):
                    dates.append(val.strip())
                else:
                    dates.append(str(val))
            
            valid_dates = [d for d in dates if d and len(d) == 10 and d[4] == '-' and d[7] == '-']
            if len(valid_dates) < 15:
                wb.close()
                os.remove(temp_path)
                return jsonify({"error": "班表格式不符：在 Row 2 第 10 欄之後找不到足夠的日期資料"}), 400
                
            # Parse agents data directly from the uploaded file
            col_mapping = {d_str: 10 + idx for idx, d_str in enumerate(dates) if d_str in valid_dates}
            agents_data = {}
            for r in range(3, sheet.max_row + 1):
                cell_a = sheet.cell(row=r, column=1).value
                name = normalize_name(cell_a)
                if name in SHIFTS_ALL:
                    default_shift = sheet.cell(row=r, column=2).value
                    meal_val = sheet.cell(row=r, column=4).value
                    
                    daily_schedule = {}
                    for d_str in valid_dates:
                        col = col_mapping[d_str]
                        cell_val = sheet.cell(row=r, column=col).value
                        daily_schedule[d_str] = cell_val
                        
                    agents_data[name] = {
                        "default_shift": default_shift,
                        "meal": meal_val,
                        "schedule": daily_schedule
                    }
            
            wb.close()
            
            # Save the parsed data directly to our in-memory schedule cache!
            # This completely avoids writing to and overwriting the original file on disk.
            schedule_cache[month_code] = {
                "dates": valid_dates,
                "agents": agents_data
            }
            
            # Clean up the temp file
            try:
                os.remove(temp_path)
            except Exception:
                pass
            
            year = month_code[:4]
            month = month_code[4:]
            agents_list = sorted(list(agents_data.keys()))
            return jsonify({
                "success": True,
                "message": f"{year}年{month}月",
                "month_code": month_code,
                "agents": agents_list
            })
            
        except Exception as e:
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except Exception:
                    pass
            return jsonify({"error": f"解析 Excel 失敗: {str(e)}"}), 400
            
    return jsonify({"error": "只支援 .xlsx 格式的 Excel 檔案"}), 400


if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5000, debug=True)
